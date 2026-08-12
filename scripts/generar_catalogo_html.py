#!/usr/bin/env python3
"""Genera landing page moderna tipo lander desde Catálogo_Final_TiendaMidu.xlsx."""
import json
from urllib.parse import quote as urlquote
from openpyxl import load_workbook

XLSX = r"D:\TIENDA DE mIDU\Catálogo_Final_TiendaMidu.xlsx"
OUT = r"D:\TIENDA DE mIDU\catalogo-midu.html"

wb = load_workbook(XLSX)
ws = wb.active
products = []
for r in range(4, ws.max_row + 1):
    d = ws.cell(row=r, column=2).value
    if not (d and isinstance(d, str) and d.startswith('D_')):
        continue
    name = (ws.cell(row=r, column=5).value or '').strip()
    sell = ws.cell(row=r, column=11).value
    offer = ws.cell(row=r, column=10).value
    if not name:
        continue
    products.append({
        'id': d, 'name': name,
        'size': (ws.cell(row=r, column=8).value or '').strip(),
        'color': (ws.cell(row=r, column=7).value or '').strip(),
        'cat': (ws.cell(row=r, column=4).value or '').strip(),
        'sell': sell, 'offer': offer,
    })

print(f"Productos leídos: {len(products)}")

# Read HTML template
TPL_PATH = "template_landing.html"
with open(TPL_PATH, "r", encoding="utf-8") as f:
    tpl = f.read()

json_data = json.dumps(products, ensure_ascii=False, indent=2)
final = tpl.replace("%%PRODUCTS_JSON%%", json_data)

with open(OUT, "w", encoding="utf-8") as f:
    f.write(final)

print(f"Guardado: {OUT}")
